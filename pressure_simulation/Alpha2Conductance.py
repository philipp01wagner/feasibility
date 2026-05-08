"""Conversion utilities for pressure/flow calibration curves.

This module mirrors the updated MATLAB reference implementation located in
``Pressure_Lypunov_Kalman_1stOrder_Demo1``. The MATLAB scripts construct
pressure, flow, and pressure-coefficient splines from experimental data stored
in ``TV_Pressure_Record.xlsx``. The Python translation keeps the same naming
conventions (``A2PSpline``, ``A2QSpline`` …) while returning a small dataclass
that bundles the spline objects together.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
from scipy.interpolate import CubicSpline
from scipy.ndimage import gaussian_filter1d


@dataclass(frozen=True)
class FlowSplines:
    """Container for the valve characteristic splines."""

    alpha_to_pressure: CubicSpline
    alpha_to_pressure_derivative: CubicSpline
    alpha_to_pressure_second_derivative: CubicSpline
    alpha_to_flow: CubicSpline
    alpha_to_flow_derivative: CubicSpline
    alpha_to_sigma: CubicSpline
    alpha_to_sigma_derivative: CubicSpline
    flow_to_alpha: CubicSpline


DEFAULT_DATA_PATH = Path("pressure_simulation/data/TV_Pressure_Record.xlsx")
_GAUSSIAN_WINDOW = 10
_INDEX_SELECT = np.array([1, 2, 3, 4, 5, 6, 7, 8, 9], dtype=int)
_Q_REFERENCE = np.array(
    [0.0, 1.5, 3.0, 4.5, 6.0, 7.5, 9.0, 10.5, 12.0, 13.5, 15.0],
    dtype=float,
)


def alpha_to_conductance(
    data_path: str | Path | None = None,
    *,
    smoothing_window: int = _GAUSSIAN_WINDOW,
) -> tuple[dict[str, float], FlowSplines]:
    """Load valve characteristics and return the system parameters plus splines."""

    target_path = Path(data_path) if data_path is not None else DEFAULT_DATA_PATH
    matrix = _load_pressure_record(target_path)
    system_params = _system_params()
    splines = _build_flow_curves_from_matrix(
        matrix=matrix,
        system_params=system_params,
        smoothing_window=smoothing_window,
    )
    return system_params, splines


def _system_params() -> dict[str, float]:
    density = 1.25  # g per standard litre
    inflow_slm = 12.0  # SLM
    qm = density * inflow_slm / 60.0  # g/s
    return {
        "R": 62.355,
        "M": 28.0,
        "qm": qm,
        "V": 34.0,
        "T": 293.15,
        "st": 1.0e-3,
        "density": density,
        "inflow_slm": inflow_slm,
    }


def _load_pressure_record(path: Path) -> np.ndarray:
    if not path.exists():
        raise FileNotFoundError(
            f"Calibration file '{path}' not found. Please provide the latest "
            "TV_Pressure_Record.xlsx exported from MATLAB."
        )

    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        data = np.genfromtxt(path, delimiter=",", dtype=float)
    elif suffix in {".xlsx", ".xls"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - import guard
            raise RuntimeError(
                "openpyxl is required to read Excel calibration files"
            ) from exc

        workbook = load_workbook(path, data_only=True, read_only=True)
        worksheet = workbook.active
        rows: list[list[float]] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            numeric_row = []
            for value in row:
                if value is None:
                    numeric_row.append(np.nan)
                else:
                    numeric_row.append(float(value))
            if all(np.isnan(numeric_row)):
                break
            rows.append(numeric_row)
        data = np.array(rows, dtype=float)
    else:  # pragma: no cover - currently unsupported formats
        raise ValueError(f"Unsupported calibration file format: {path.suffix}")

    if data.ndim != 2 or data.shape[1] < 3:
        raise ValueError("Calibration file must contain alpha and >=2 pressure columns")

    mask_rows = ~np.all(np.isnan(data), axis=1)
    data = data[mask_rows]
    if data.size == 0:
        raise ValueError("Calibration file does not contain usable data")

    # Trim columns that are entirely NaN (e.g., padding in the spreadsheet)
    mask_cols = ~np.all(np.isnan(data), axis=0)
    return data[:, mask_cols]


def _build_flow_curves_from_matrix(
    *,
    matrix: np.ndarray,
    system_params: dict[str, float],
    smoothing_window: int,
) -> FlowSplines:
    alpha_raw = matrix[:, 0]
    measurements = matrix[:, 1:]

    if measurements.shape[1] <= np.max(_INDEX_SELECT):
        raise ValueError(
            "Calibration data must provide at least 10 flow columns to "
            "compute Q and sigma."
        )

    # Remove NaNs via linear interpolation per column
    valid_mask = ~np.isnan(alpha_raw)
    alpha_raw = alpha_raw[valid_mask]
    measurements = measurements[valid_mask]

    for col in range(measurements.shape[1]):
        column = measurements[:, col]
        if np.isnan(column).any():
            valid = ~np.isnan(column)
            measurements[:, col] = np.interp(
                alpha_raw,
                alpha_raw[valid],
                column[valid],
            )

    alpha_smooth = np.linspace(float(np.min(alpha_raw)), float(np.max(alpha_raw)), 51)

    selected = measurements[:, _INDEX_SELECT]
    p_smooth = np.empty((alpha_smooth.size, selected.shape[1]))
    for idx in range(selected.shape[1]):
        spline = CubicSpline(alpha_raw, selected[:, idx])
        interpolated = spline(alpha_smooth)
        if smoothing_window > 1:
            sigma = max((smoothing_window - 1) / 6.0, 1e-6)
            interpolated = gaussian_filter1d(interpolated, sigma=sigma, mode="nearest")
        p_smooth[:, idx] = interpolated

    R = system_params["R"]
    T = system_params["T"]
    M = system_params["M"]

    q_mass = (
        _Q_REFERENCE[_INDEX_SELECT] * system_params["density"] / 60.0
    )  # g/s for each calibration column

    Q = np.empty(alpha_smooth.size)
    Qsigma = np.empty(alpha_smooth.size)
    sqrt_q = np.sqrt(q_mass)
    rhs = (R * T / M) * q_mass[[0, 7]]

    for i, _alpha in enumerate(alpha_smooth):
        a_matrix = np.array(
            [
                [p_smooth[i, 0], -sqrt_q[0]],
                [p_smooth[i, 7], -sqrt_q[7]],
            ],
            dtype=float,
        )
        solution = np.linalg.solve(a_matrix, rhs)
        Q[i] = solution[0]
        Qsigma[i] = solution[1]

    sigma_vals = Qsigma / Q

    alpha_to_pressure = CubicSpline(alpha_smooth, p_smooth[:, 6])
    alpha_to_flow = CubicSpline(alpha_smooth, Q)
    alpha_to_sigma = CubicSpline(alpha_smooth, sigma_vals)

    # Derivatives
    a2dpda = alpha_to_pressure.derivative(1)
    a2d2pda2 = alpha_to_pressure.derivative(2)
    a2dqda = alpha_to_flow.derivative(1)
    a2dsigma = alpha_to_sigma.derivative(1)

    sort_idx = np.argsort(Q)
    flow_to_alpha = CubicSpline(Q[sort_idx], alpha_smooth[sort_idx])

    return FlowSplines(
        alpha_to_pressure=alpha_to_pressure,
        alpha_to_pressure_derivative=a2dpda,
        alpha_to_pressure_second_derivative=a2d2pda2,
        alpha_to_flow=alpha_to_flow,
        alpha_to_flow_derivative=a2dqda,
        alpha_to_sigma=alpha_to_sigma,
        alpha_to_sigma_derivative=a2dsigma,
        flow_to_alpha=flow_to_alpha,
    )


if __name__ == "__main__":  # pragma: no cover - manual smoke test helper
    params, splines = alpha_to_conductance()
    print("Loaded system parameters:", params)
    print(
        "Pressure at 50%:",
        float(splines.alpha_to_pressure(50.0)),
    )
    print(
        "Outflow coefficient at 50%:",
        float(splines.alpha_to_flow(50.0)),
    )
